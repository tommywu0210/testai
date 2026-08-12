import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook

# ─────────────────────────────────────────────
# 配置
# ─────────────────────────────────────────────
EXCEL_PATH = r"C:\Users\I0438687\OneDrive - Sanofi\Desktop\Demo.xlsx"
SHEET_NAME = "Table"
REMARK_COL = "备注"

st.set_page_config(
    page_title="数据展示小程序",
    page_icon="📊",
    layout="wide"
)

# ─────────────────────────────────────────────
# 读取 Excel
# ─────────────────────────────────────────────
def load_data():
    if not os.path.exists(EXCEL_PATH):
        st.error(f"❌ 找不到文件：{EXCEL_PATH}\n请确认文件路径是否正确。")
        st.stop()
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, dtype=str)
    if REMARK_COL not in df.columns:
        df[REMARK_COL] = ""
    df[REMARK_COL] = df[REMARK_COL].fillna("")
    return df

# ─────────────────────────────────────────────
# 保存备注回写 Excel（只更新备注列，保留原有数据格式）
# ─────────────────────────────────────────────
def save_remarks(remarks: list):
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 找到或新建备注列
    header_row = [cell.value for cell in ws[1]]
    if REMARK_COL in header_row:
        remark_col_idx = header_row.index(REMARK_COL) + 1
    else:
        remark_col_idx = len(header_row) + 1
        ws.cell(row=1, column=remark_col_idx, value=REMARK_COL)

    # 写入每行备注（从第2行开始，第1行是表头）
    for i, remark in enumerate(remarks):
        ws.cell(row=i + 2, column=remark_col_idx, value=remark)

    wb.save(EXCEL_PATH)

# ─────────────────────────────────────────────
# 主界面
# ─────────────────────────────────────────────
st.title("📊 数据展示小程序")
st.caption(f"数据来源：{EXCEL_PATH}")

# 刷新按钮
col_refresh, col_spacer = st.columns([1, 9])
with col_refresh:
    if st.button("🔄 刷新数据"):
        st.cache_data.clear()
        st.rerun()

# 加载数据
df = load_data()
data_cols = [c for c in df.columns if c != REMARK_COL]

st.markdown("---")
st.subheader(f"共 {len(df)} 条记录")

# ─────────────────────────────────────────────
# 表格 + 备注输入（按行展示）
# ─────────────────────────────────────────────
# 表头
header_cols = st.columns([1] * len(data_cols) + [2, 0.5])
for i, col_name in enumerate(data_cols):
    header_cols[i].markdown(f"**{col_name}**")
header_cols[-2].markdown(f"**{REMARK_COL}**")
header_cols[-1].markdown("")

st.markdown("---")

# 初始化 session_state 存储备注
if "remarks" not in st.session_state:
    st.session_state.remarks = df[REMARK_COL].tolist()

# 每行数据
for idx, row in df.iterrows():
    row_cols = st.columns([1] * len(data_cols) + [2, 0.5])
    
    # 显示数据列
    for i, col_name in enumerate(data_cols):
        row_cols[i].write(row[col_name])
    
    # 备注输入框
    remark_val = st.session_state.remarks[idx] if idx < len(st.session_state.remarks) else ""
    new_remark = row_cols[-2].text_input(
        label=f"备注_{idx}",
        value=remark_val,
        label_visibility="collapsed",
        placeholder="请输入备注...",
        key=f"remark_{idx}"
    )
    st.session_state.remarks[idx] = new_remark

st.markdown("---")

# ─────────────────────────────────────────────
# 保存按钮
# ─────────────────────────────────────────────
col_save, col_msg = st.columns([1, 5])
with col_save:
    if st.button("💾 保存备注", type="primary", use_container_width=True):
        try:
            remarks_to_save = [st.session_state.get(f"remark_{i}", "") for i in range(len(df))]
            save_remarks(remarks_to_save)
            col_msg.success("✅ 备注已成功保存到 Excel 文件！")
        except Exception as e:
            col_msg.error(f"❌ 保存失败：{e}")

st.markdown("---")
st.caption("💡 提示:修改备注后点击【保存备注】按钮，备注将自动写入 Excel 文件，下次打开时自动显示。")